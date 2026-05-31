import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from fpdf import FPDF
from src.models.data_models import SchoolSystemModel

class Exporter:
    def __init__(self, model: SchoolSystemModel):
        self.model = model

    def export_excel_grupos(self, filepath: str):
        wb = Workbook()
        wb.remove(wb.active) # Remove default sheet
        
        for grupo, horario in self.model.horarios.items():
            ws = wb.create_sheet(title=f"Grupo {grupo}")
            self._write_excel_sheet(ws, grupo, horario, self.model.dias, self.model.horas, self.model.indices_recreo)
            
        wb.save(filepath)

    def export_excel_docentes(self, filepath: str):
        wb = Workbook()
        wb.remove(wb.active)
        
        # Build teacher schedules
        docentes_horario = {}
        for grupo, horario in self.model.horarios.items():
            for (d, h), clase in horario.items():
                if clase.docente not in docentes_horario:
                    docentes_horario[clase.docente] = {}
                docentes_horario[clase.docente][(d, h)] = f"{clase.materia}\n({grupo})"
                
        for docente, horario in docentes_horario.items():
            # Clean sheet name characters if necessary
            ws = wb.create_sheet(title=str(docente)[:31])
            self._write_excel_sheet(ws, f"Horario de {docente}", horario, self.model.dias, self.model.horas, self.model.indices_recreo, is_docente=True)
            
        if not wb.sheetnames:
            wb.create_sheet("Vacio")
            
        wb.save(filepath)

    def _write_excel_sheet(self, ws, titulo, horario, dias, horas, recreos, is_docente=False):
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        recreo_fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")
        cell_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(dias)+1)
        ws.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=14)
        ws.cell(row=1, column=1).alignment = center_align

        # Headers
        ws.cell(row=2, column=1, value="HORA").font = header_font
        ws.cell(row=2, column=1).fill = header_fill
        ws.cell(row=2, column=1).alignment = center_align
        ws.column_dimensions['A'].width = 15

        for i, d in enumerate(dias):
            col = i + 2
            cell = ws.cell(row=2, column=col, value=d.upper())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            ws.column_dimensions[cell.column_letter].width = 20

        # Data
        for r, h in enumerate(horas):
            row_idx = r + 3
            ws.cell(row=row_idx, column=1, value=h).alignment = center_align
            ws.cell(row=row_idx, column=1).border = thin_border
            
            es_rec = r in recreos
            for c, d in enumerate(dias):
                col_idx = c + 2
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = center_align
                
                if es_rec:
                    cell.value = "--- RECREO ---"
                    cell.fill = recreo_fill
                else:
                    item = horario.get((c, r))
                    if item:
                        if is_docente:
                            cell.value = item # It's a string for teacher view
                        else:
                            cell.value = f"{item.materia}\n{item.docente}"
                        cell.fill = PatternFill(start_color=item.color_bg.replace("#","") if not is_docente else "3498DB", end_color=item.color_bg.replace("#","") if not is_docente else "3498DB", fill_type="solid")
                        cell.font = Font(color="FFFFFF")
                    else:
                        cell.fill = cell_fill

    def export_pdf_grupos(self, filepath: str):
        pdf = FPDF(orientation='L', unit='mm', format='A4')
        pdf.set_auto_page_break(auto=True, margin=15)
        
        for grupo, horario in self.model.horarios.items():
            pdf.add_page()
            pdf.set_font("helvetica", 'B', 16)
            pdf.cell(0, 10, f"Escuela: {self.model.nombre_escuela} - Ciclo: {self.model.ciclo_escolar}", new_x="LMARGIN", new_y="NEXT", align='C')
            pdf.set_font("helvetica", 'B', 14)
            pdf.cell(0, 10, f"Horario del Grupo: {grupo}", new_x="LMARGIN", new_y="NEXT", align='C')
            pdf.ln(5)
            
            col_w = 250 / (len(self.model.dias) + 1)
            row_h = 15
            
            # Header
            pdf.set_font("helvetica", 'B', 10)
            pdf.set_fill_color(44, 62, 80)
            pdf.set_text_color(255, 255, 255)
            
            pdf.cell(col_w, row_h, "HORA", border=1, align='C', fill=True)
            for d in self.model.dias:
                pdf.cell(col_w, row_h, d.upper(), border=1, align='C', fill=True)
            pdf.ln(row_h)
            
            # Body
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("helvetica", '', 9)
            
            for r, h in enumerate(self.model.horas):
                pdf.set_fill_color(236, 240, 241)
                es_rec = r in self.model.indices_recreo
                
                x_start = pdf.get_x()
                y_start = pdf.get_y()
                
                pdf.set_xy(x_start, y_start)
                pdf.cell(col_w, row_h, h, border=1, align='C', fill=True)
                
                for c, d in enumerate(self.model.dias):
                    x_curr = x_start + (c + 1) * col_w
                    pdf.set_xy(x_curr, y_start)
                    
                    if es_rec:
                        pdf.set_fill_color(149, 165, 166)
                        pdf.cell(col_w, row_h, "--- RECREO ---", border=1, align='C', fill=True)
                    else:
                        item = horario.get((c, r))
                        if item:
                            pdf.set_fill_color(52, 152, 219)
                            pdf.set_text_color(255, 255, 255)
                            
                            # Draw cell background and border
                            pdf.cell(col_w, row_h, "", border=1, fill=True)
                            
                            # Write text on two lines
                            pdf.set_xy(x_curr, y_start + 2)
                            pdf.cell(col_w, row_h/2 - 2, item.materia[:20], align='C')
                            pdf.set_xy(x_curr, y_start + row_h/2)
                            pdf.cell(col_w, row_h/2 - 2, item.docente[:20], align='C')
                            
                            # Restore cursor to end of cell
                            pdf.set_xy(x_curr + col_w, y_start)
                            pdf.set_text_color(0, 0, 0)
                        else:
                            pdf.set_fill_color(255, 255, 255)
                            pdf.cell(col_w, row_h, "", border=1, align='C', fill=True)
                            
                pdf.ln(row_h)
                
        pdf.output(filepath)
